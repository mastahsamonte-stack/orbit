#!/usr/bin/env python3
"""Builds the DSP/HCS Staffing Agency financial model with live formulas.

Robust version: every row's actual cell address is tracked in a dict so formulas
reference cells by NAME, not by hand-counted row numbers.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

wb = openpyxl.Workbook()

H1 = Font(bold=True, size=16, color="FFFFFF")
H2 = Font(bold=True, size=12, color="1F3864")
BOLD = Font(bold=True)
WHITE_BOLD = Font(bold=True, color="FFFFFF")
ITAL = Font(italic=True, size=9, color="808080")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
CALC_FILL = PatternFill("solid", fgColor="E2EFDA")
HDR_FILL = PatternFill("solid", fgColor="1F3864")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

USD = '$#,##0'; USD2 = '$#,##0.00'; PCT = '0.0%'; MULT = '0.0"x"'; NUM = '#,##0'


class Sheet:
    """Helper that appends rows top-to-bottom and remembers each row's address."""
    def __init__(self, ws, title, w0=46, w1=18, w2=None):
        self.ws = ws; self.r = 1; self.rows = {}
        ws.column_dimensions['A'].width = w0
        ws.column_dimensions['B'].width = w1
        if w2: ws.column_dimensions['C'].width = w2
        # title bar
        c = ws.cell(row=1, column=1, value=title); c.font = H1; c.fill = HDR_FILL
        for col in range(1, (3 if w2 else 2) + 1):
            ws.cell(row=1, column=col).fill = HDR_FILL
        ws.row_dimensions[1].height = 24
        self.r = 2

    def section(self, text, span=2):
        c = self.ws.cell(row=self.r, column=1, value=text); c.font = H2; c.fill = SUB_FILL
        for col in range(1, span + 1):
            self.ws.cell(row=self.r, column=col).fill = SUB_FILL
        self.r += 1

    def blank(self):
        self.r += 1

    def note(self, text):
        c = self.ws.cell(row=self.r, column=1, value=text); c.font = ITAL
        self.r += 1

    def addr(self, key):
        return f"'{self.ws.title}'!B{self.rows[key]}"

    def line(self, key, label, value=None, formula=None, fmt=None, bold=False,
             note=None, is_input=False):
        self.ws.cell(row=self.r, column=1, value=label)
        cell = self.ws.cell(row=self.r, column=2)
        if formula is not None:
            cell.value = formula; cell.fill = CALC_FILL
        else:
            cell.value = value; cell.fill = INPUT_FILL if is_input else CALC_FILL
        cell.border = BORDER
        if fmt: cell.number_format = fmt
        if bold: cell.font = BOLD
        if note:
            n = self.ws.cell(row=self.r, column=3, value=note); n.font = ITAL
        self.rows[key] = self.r
        self.r += 1
        return cell


# =====================================================================
# 1. README
# =====================================================================
ws = wb.active; ws.title = "README"; ws.column_dimensions['A'].width = 102
c = ws.cell(row=1, column=1, value="DSP / HCS Staffing Agency — Financial Model")
c.font = H1; c.fill = HDR_FILL; ws.row_dimensions[1].height = 24
readme = [
    "", "HOW TO USE",
    "  - YELLOW cells = inputs you edit.  GREEN cells = calculated (don't overwrite).",
    "  - Change any yellow input and all dependent green cells recalculate automatically.",
    "  - Works in Excel or Google Sheets. Formulas are live.",
    "", "TABS",
    "  1. Assumptions     - all inputs in one place.",
    "  2. Unit Economics  - bill vs. pay rate, gross margin per hour, annual P&L.",
    "  3. Working Capital - payroll float: cash gap between paying DSPs and getting paid.",
    "  4. Acquisition     - pro-forma for buying an agency (multiple, price, funding, returns).",
    "", "GROUNDING (Texas / San Antonio, 2025-2026 - verify before relying)",
    "  - TX Medicaid attendant rates support avg ~$13.00/hr wage + 14-15% payroll-tax/benefits (Sept 2025).",
    "  - On Medicaid work, margin comes from overhead efficiency & scale, NOT rate negotiation (capped).",
    "  - Staffing markup typically 1.45x-1.75x of pay; gross margin ~15-30%.",
    "  - Payroll float: pay DSPs weekly, collect from Medicaid/MCOs in 30-90 days. Fund it or factor.",
    "  - Small agencies ($0.5-3M rev) trade ~2-3.5x SDE; IDD/HCS ~4-7x EBITDA (small), 8-11x (platform).",
    "  - Caregiver turnover median ~75% (2024) - the #1 operational risk.",
    "", "These are starting assumptions, not guarantees. Replace with real diligence numbers.",
]
r = 2
for line in readme:
    cc = ws.cell(row=r, column=1, value=line)
    if line.strip() and line.strip() == line.strip().upper() and not line.startswith("  "):
        cc.font = H2
    r += 1

# =====================================================================
# 2. ASSUMPTIONS
# =====================================================================
wa = Sheet(wb.create_sheet("Assumptions"), "Assumptions  (edit YELLOW cells)", 42, 16, 52)
# header row for the note column
wa.ws.cell(row=2, column=3, value="Note").font = WHITE_BOLD
wa.ws.cell(row=2, column=3).fill = SUB_FILL
wa.ws.cell(row=2, column=1, value="Input").font = BOLD
wa.ws.cell(row=2, column=2, value="Value").font = BOLD
wa.r = 3

def ainp(key, label, val, fmt, note):
    wa.line(key, label, value=val, fmt=fmt, is_input=True, note=note)

wa.section("LABOR & BILLING", span=3)
ainp("pay", "DSP pay rate ($/hr)", 14.00, USD2, "What you pay the caregiver (~$13 Medicaid avg; pay up to compete).")
ainp("mult", "Bill rate multiplier (x pay)", 1.60, MULT, "1.45-1.75 typical. Medicaid is rate-capped - confirm allowable.")
ainp("hrs_wk", "Billable hours per DSP per week", 32, NUM, "Net of unfilled shifts, PTO, no-shows.")
ainp("dsps", "Number of DSPs (FTE-equiv)", 40, NUM, "Drives total volume.")
ainp("weeks", "Weeks operating per year", 52, NUM, "")
wa.blank()
wa.section("LABOR BURDEN (on top of pay)", span=3)
ainp("tax", "Employer payroll taxes (FICA+FUTA+SUTA) %", 0.10, PCT, "~8-12% of wages.")
ainp("wc", "Workers' comp %", 0.05, PCT, "5-15%+ for caregiving. Optional in TX but recommended.")
ainp("ben", "Benefits / PTO %", 0.04, PCT, "Raise to compete on retention.")
wa.blank()
wa.section("OVERHEAD (fixed, annual)", span=3)
ainp("rent", "Office rent + utilities", 36000, USD, "")
ainp("admin", "Admin & office salaries (non-billable)", 180000, USD, "Administrator, scheduler, biller, recruiter.")
ainp("ins", "Insurance (GL + professional + EPLI)", 12000, USD, "")
ainp("sw", "Software (EVV, scheduling, payroll)", 9000, USD, "HHAeXchange EVV free for Medicaid; other tools paid.")
ainp("comp", "Compliance, training, licensing", 8000, USD, "HCSSA $2,625/3yr, admin training, background checks.")
ainp("mktg", "Marketing & recruiting", 24000, USD, "Recruiting DSPs is the real bottleneck.")
ainp("ga", "Other G&A", 18000, USD, "")
wa.blank()
wa.section("WORKING CAPITAL", span=3)
ainp("dso", "Days to collect (Medicaid/MCO)", 45, NUM, "30-90 days typical for Medicaid.")
ainp("payfreq", "Payroll frequency (days)", 7, NUM, "Weekly payroll = bigger float.")
ainp("factor_on", "Factoring? (1=yes, 0=no)", 0, NUM, "If 1, applies factoring lines below.")
ainp("adv", "Factoring advance rate", 0.85, PCT, "70-92% typical.")
ainp("fcost", "Factoring annual cost rate", 0.24, PCT, "~18-36% annualized.")
wa.blank()
wa.section("ACQUISITION", span=3)
ainp("t_rev", "Target trailing revenue", 1500000, USD, "From seller financials.")
ainp("t_marg", "Target EBITDA margin", 0.15, PCT, "Normalized after market-rate owner salary.")
ainp("t_mult", "Purchase multiple (x EBITDA)", 5.0, MULT, "Small IDD ~4-7x; platforms 8-11x.")
ainp("debt_pct", "% financed with debt", 0.50, PCT, "SBA / seller note / lender.")
ainp("rate", "Debt interest rate", 0.105, PCT, "Illustrative.")
ainp("raise", "Equity raise target", 750000, USD, "What you raise from investors.")

def A(key):
    return wa.addr(key)

# =====================================================================
# 3. UNIT ECONOMICS
# =====================================================================
ue = Sheet(wb.create_sheet("Unit Economics"), "Unit Economics", 44, 18)
ue.section("PER BILLED HOUR")
ue.line("pay", "DSP pay rate", formula=f"={A('pay')}", fmt=USD2)
ue.line("bill", "Bill rate", formula=f"={A('pay')}*{A('mult')}", fmt=USD2)
ue.line("burden", "Labor burden $/hr", formula=f"={A('pay')}*({A('tax')}+{A('wc')}+{A('ben')})", fmt=USD2)
ue.line("loaded", "Fully loaded labor cost $/hr", formula=f"={ue.addr('pay')}+{ue.addr('burden')}", fmt=USD2)
ue.line("gp_hr", "Gross profit $/hr", formula=f"={ue.addr('bill')}-{ue.addr('loaded')}", fmt=USD2, bold=True)
ue.line("gm_hr", "Gross margin %", formula=f"={ue.addr('gp_hr')}/{ue.addr('bill')}", fmt=PCT)
ue.blank()
ue.section("ANNUAL VOLUME")
ue.line("hours", "Total billable hours / yr", formula=f"={A('hrs_wk')}*{A('dsps')}*{A('weeks')}", fmt=NUM)
ue.blank()
ue.section("ANNUAL P&L")
ue.line("rev", "Revenue", formula=f"={ue.addr('bill')}*{ue.addr('hours')}", fmt=USD, bold=True)
ue.line("cogs", "Cost of labor (pay + burden)", formula=f"={ue.addr('loaded')}*{ue.addr('hours')}", fmt=USD)
ue.line("gp", "Gross profit", formula=f"={ue.addr('rev')}-{ue.addr('cogs')}", fmt=USD, bold=True)
ue.line("gm", "Gross margin %", formula=f"={ue.addr('gp')}/{ue.addr('rev')}", fmt=PCT)
ue.line("oh", "Total overhead (fixed)",
        formula=f"=SUM({A('rent')},{A('admin')},{A('ins')},{A('sw')},{A('comp')},{A('mktg')},{A('ga')})", fmt=USD)
ue.line("ebitda", "EBITDA", formula=f"={ue.addr('gp')}-{ue.addr('oh')}", fmt=USD, bold=True)
ue.line("ebitda_m", "EBITDA margin %", formula=f"={ue.addr('ebitda')}/{ue.addr('rev')}", fmt=PCT)

# =====================================================================
# 4. WORKING CAPITAL
# =====================================================================
wc = Sheet(wb.create_sheet("Working Capital"), "Working Capital - Payroll Float", 46, 18)
wc.line("rev", "Annual revenue", formula=f"={ue.addr('rev')}", fmt=USD)
wc.line("labor", "Annual labor cost (cash out)", formula=f"={ue.addr('cogs')}", fmt=USD)
wc.line("drev", "Avg daily revenue", formula=f"={wc.addr('rev')}/365", fmt=USD)
wc.line("dlab", "Avg daily labor cost", formula=f"={wc.addr('labor')}/365", fmt=USD)
wc.blank()
wc.line("dso", "Days to collect from payer", formula=f"={A('dso')}", fmt=NUM)
wc.line("pay", "Payroll cycle (days)", formula=f"={A('payfreq')}", fmt=NUM)
wc.line("gap", "Float gap (days)", formula=f"={wc.addr('dso')}-{wc.addr('pay')}", fmt=NUM)
wc.blank()
wc.section("PEAK WORKING CAPITAL NEEDED")
wc.line("ar", "AR outstanding (revenue in transit)", formula=f"={wc.addr('drev')}*{wc.addr('dso')}", fmt=USD)
wc.line("laborgap", "Labor cash paid out during gap", formula=f"={wc.addr('dlab')}*{wc.addr('gap')}", fmt=USD)
wc.line("need", "Est. working capital required", formula=f"=MAX({wc.addr('ar')},{wc.addr('laborgap')})", fmt=USD, bold=True)
wc.blank()
wc.section("IF FACTORING (toggle on Assumptions)")
wc.line("fadv", "Cash advanced upfront", formula=f"={wc.addr('ar')}*{A('adv')}*{A('factor_on')}", fmt=USD)
wc.line("fcost", "Annual factoring cost", formula=f"={wc.addr('rev')}*{A('fcost')}*{A('factor_on')}", fmt=USD)

# =====================================================================
# 5. ACQUISITION
# =====================================================================
aq = Sheet(wb.create_sheet("Acquisition"), "Acquisition Pro-Forma", 46, 18)
aq.section("TARGET")
aq.line("rev", "Trailing revenue", formula=f"={A('t_rev')}", fmt=USD)
aq.line("marg", "EBITDA margin", formula=f"={A('t_marg')}", fmt=PCT)
aq.line("ebitda", "EBITDA", formula=f"={aq.addr('rev')}*{aq.addr('marg')}", fmt=USD, bold=True)
aq.blank()
aq.section("PURCHASE")
aq.line("mult", "Purchase multiple", formula=f"={A('t_mult')}", fmt=MULT)
aq.line("price", "Enterprise value (purchase price)", formula=f"={aq.addr('ebitda')}*{aq.addr('mult')}", fmt=USD, bold=True)
aq.blank()
aq.section("FUNDING")
aq.line("debt", "Debt financed", formula=f"={aq.addr('price')}*{A('debt_pct')}", fmt=USD)
aq.line("equity", "Equity required", formula=f"={aq.addr('price')}-{aq.addr('debt')}", fmt=USD, bold=True)
aq.line("raise", "Equity raise target", formula=f"={A('raise')}", fmt=USD)
aq.line("surplus", "Surplus / (shortfall) vs. raise", formula=f"={aq.addr('raise')}-{aq.addr('equity')}", fmt=USD, bold=True)
aq.blank()
aq.section("RETURNS (illustrative, pre-tax, year 1, interest-only)")
aq.line("ds", "Annual debt service (interest only)", formula=f"={aq.addr('debt')}*{A('rate')}", fmt=USD)
aq.line("cf", "EBITDA after debt service", formula=f"={aq.addr('ebitda')}-{aq.addr('ds')}", fmt=USD, bold=True)
aq.line("coc", "Cash-on-cash return on equity", formula=f"={aq.addr('cf')}/{aq.addr('equity')}", fmt=PCT, bold=True)
aq.line("payback", "Implied payback (yrs on equity)", formula=f"={aq.addr('equity')}/{aq.addr('cf')}", fmt='0.0')
aq.blank()
aq.note("Simplified: interest-only, pre-tax, single year. For a raise, build a 5-yr projection with")
aq.note("census ramp, principal amortization, and taxes.")

wb.save("/home/user/orbit/staffing-agency-plan/Staffing_Agency_Financial_Model.xlsx")
print("saved")
